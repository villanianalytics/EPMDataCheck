import hashlib
from io import StringIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from app.core.logging_engine import *


def ensure_unique_column_names(headers):
    """
    Ensure each column name is unique by appending an index to duplicates.
    """
    unique_headers = []
    seen = {}
    for header in headers:
        new_header = header
        count = seen.get(header, 0)
        while new_header in seen:
            count += 1
            new_header = f"{header}_{count}"
        seen[header] = count
        unique_headers.append(new_header)
    return unique_headers


def read_and_process_csv(csv_content, header_rows=2, row_dimensions=[]):
    lines = csv_content.split('\n')
    header_lines = lines[:header_rows]
    data_lines = lines[header_rows:]

    header_str = '\n'.join(header_lines)
    data_str = '\n'.join(data_lines)

    header_df = pd.read_csv(StringIO(header_str), header=None)
    concatenated_headers = ['_'.join(filter(None, header_df[col].astype(str))) for col in header_df.columns]

    # Assign row dimensions to the first N headers, then ensure all column names are unique
    for i, dimension in enumerate(row_dimensions):
        if i < len(concatenated_headers):
            concatenated_headers[i] = dimension
    concatenated_headers = ensure_unique_column_names(concatenated_headers)

    df = pd.read_csv(StringIO(data_str), names=concatenated_headers)
    df = df.apply(pd.to_numeric, errors='ignore')

    return df


def check_row_failure(row):
    # Check if any variance column in the row has a failure (not equal to 0 in this example)
    for col in row.index:
        if "_variance" in col and row[col] != 0:
            return "fail"
    return "success"


def save_to_excel_with_hash_check(data_pull_csv, comparison_csv, excel_path, src_num_headers, tgt_num_headers,
                                  row_dimensions):
    # Process source and target CSVs
    df_source = read_and_process_csv(data_pull_csv, src_num_headers, row_dimensions)
    df_target = read_and_process_csv(comparison_csv, tgt_num_headers, row_dimensions)

    # Perform hash check for exact match
    hash_pull = hashlib.sha256(df_source.to_string().encode()).hexdigest()
    hash_comp = hashlib.sha256(df_target.to_string().encode()).hexdigest()
    match = hash_pull == hash_comp

    # Set row_dimensions as index for comparison
    df_source.set_index(row_dimensions, inplace=True)
    df_target.set_index(row_dimensions, inplace=True)

    comparison_df = df_source.join(df_target, how='outer', lsuffix='_source', rsuffix='_target')

    for col in set([c.rsplit('_', 1)[0] for c in comparison_df.columns if '_source' in c]):
        source_col = f'{col}_source'
        target_col = f'{col}_target'
        variance_col = f'{col}_variance'
        notes_col = f'{col}_notes'

        comparison_df[variance_col] = comparison_df[source_col].sub(comparison_df[target_col], fill_value=0)
        comparison_df[notes_col] = comparison_df.apply(
            lambda row: 'Data Match' if row[source_col] == row[target_col]
            else 'Values do not match' if not pd.isna(row[source_col]) and not pd.isna(row[target_col])
            else 'Data exists in source, not in target' if pd.isna(row[target_col])
            else 'Data exists in target, not in source', axis=1
        )

        # Apply the check_row_failure function across the DataFrame rows
        # Axis=1 specifies that the function should be applied to each row, not column
        comparison_df['rowfailure'] = comparison_df.apply(check_row_failure, axis=1)

    comparison_df.reset_index(inplace=True)

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_source.reset_index().to_excel(writer, sheet_name='Source', index=False)
        df_target.reset_index().to_excel(writer, sheet_name='Target', index=False)
        comparison_df.to_excel(writer, sheet_name='Validation', index=False)

    # Load the workbook and select the sheet
    workbook = load_workbook(excel_path)
    sheet = workbook['Validation']

    # Define fills for conditional formatting
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Dynamically identify variance and notes columns based on headers
    variance_columns = []
    notes_columns = []
    failure_column =[]

    for col in range(1, sheet.max_column + 1):  # Iterate through columns
        header_value = sheet.cell(row=1, column=col).value
        if header_value and "_variance" in header_value:  # Adjust the condition based on your naming pattern
            variance_columns.append(get_column_letter(col))
        elif header_value and "_notes" in header_value:  # Adjust the condition based on your naming pattern
            notes_columns.append(get_column_letter(col))
        elif header_value and "rowfailure" in header_value:  # Adjust the condition based on your naming pattern
            failure_column.append(get_column_letter(col))

    # Apply conditional formatting to identified variance columns
    for col in variance_columns:
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='notEqual', formula=['0'], fill=red_fill))
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='equal', formula=['0'], fill=green_fill))

    # Apply conditional formatting to identified notes columns
    for col in notes_columns:
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='notEqual', formula=['"Data Match"'], fill=red_fill))
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='equal', formula=['"Data Match"'], fill=green_fill))

    for col in failure_column:
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='notEqual', formula=['"success"'], fill=red_fill))
        sheet.conditional_formatting.add(f'{col}2:{col}{sheet.max_row}',
                                         CellIsRule(operator='equal', formula=['"success"'], fill=green_fill))

    for row in range(2, sheet.max_row + 1):  # Start from 2 to skip the header
        rowfailure_cell_value = sheet[f'{failure_column[0]}{row}'].value
        if rowfailure_cell_value == "success":
            sheet.row_dimensions[row].hidden = True

    # Since you want to freeze at #Row_Dims+1, #header_dims+1, for 2 dimensions and 2 headers, it's C3
    freeze_cell = f"{get_column_letter(len(row_dimensions) + 1)}{2}"

    # Set the freeze panes
    sheet.freeze_panes = freeze_cell
    # Save the workbook with conditional formatting
    workbook.save(excel_path)

    print('Excel file created with three tabs: Source, Target, and Validation.')
    return 200 if match else 412
